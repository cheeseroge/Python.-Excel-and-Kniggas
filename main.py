import openpyxl, random
from openpyxl import Workbook
from openpyxl.styles import Font

wb1 = openpyxl.load_workbook('Knigga1.xlsx')
ws0 = wb1.active

wb = Workbook()
ws = wb.active

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 7
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 21

ws.title="randomnoe zapolnenie"

x1 = ws.cell(row=1, column=1)
x1.value = "ФИО"
x2 = ws.cell(row=1, column=2)
x2.value = "Сумма"
x3 = ws.cell(row=1, column=3)
x3.value = "Месяц последней сдачи"
x4 = ws.cell(row=1, column=4)
x4.value = "Взносов достаточно?"

for i in range(0, 30):
    ws['A' + str(i + 2)] = ws0['A' + str(i + 2)].value

for i in range(0,29):
    month = ['Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май']
    ws['C'+str(i+2)]=random.choice(month)

for i in range(0, 29):
    ws['B' + str(i + 2)] = random.randint(0, 5500)

for i in range(0, 29):
    if ws['B' + str(i + 2)].value >= 2500:
        ws['D' + str(i + 2)] = '+'
    else:
        ws['D' + str(i + 2)] = '-'

wb.save('Knigga2.xlsx')

wb1 = openpyxl.load_workbook('Knigga2.xlsx')
ws2 = wb1.active

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 35

ws.title="resultat"

y1 = ws.cell(row=1, column=1)
y1.value = "ФИО"
y2 = ws.cell(row=1, column=2)
y2.value = "Место отдыха"
y3 = ws.cell(row=1, column=3)
y3.value = "Статус"
y4 = ws.cell(row=1, column=4)
y4.value = " "

for i in range(0, 30):
    ws['A' + str(i + 2)] = ws2['A' + str(i + 2)].value
lager = ['Петушки', 'Бауманец', 'Корпус Энерго', 'Артек', 'Орлёнок', 'Смена', 'Магадан', 'Трудовой лагерь', 'Интеграл', 'Лагерь Деда Мотиса']
for i in range(0, 29):
    ws['B' + str(i + 2)] = random.choice(lager)
for i in range(0, 29):
    for i in range(0, 29):
        if (ws['D' + str(i + 2)].value == '+'):
            ws['C' + str(i + 2)] = 'Едет'
        else:
            ws['C' + str(i + 2)] = 'Не едет'
for i in range(0, 29):
    ws['D' + str(i + 2)] = " "

wb.save('Knigga3.xlsx')